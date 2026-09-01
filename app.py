import streamlit as st
import pandas as pd
import re
import os
import pdfplumber
from io import BytesIO
import openpyxl
from openpyxl.styles import Alignment

# Configure the web page
st.set_page_config(page_title="SRM Schedule Consolidator", page_icon="🤖", layout="centered")
st.title("SRM Schedule PDF Consolidator 🤖 (v6.0)") 
st.write("Drag and drop your PDF schedules below to instantly generate your formatted Excel sheet.")

uploaded_files = st.file_uploader("Upload PDF files", type="pdf", accept_multiple_files=True)

if uploaded_files:
    if st.button("Process Files"):
        all_rows = []
        
        with st.spinner("Extracting with Deep Structural Table Analysis..."):
            for file in uploaded_files:
                file_name = file.name
                try:
                    file_bytes = file.getvalue()
                    
                    # Using pdfplumber to rip the table structure cell-by-cell
                    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
                        for page_num, page in enumerate(pdf.pages):
                            try:
                                # 1. Extract standard surface text
                                page_text = page.extract_text() or ""
                                
                                # 2. Extract deep grid text (Bypasses table invisibility)
                                tables = page.extract_tables()
                                table_text = ""
                                for table in tables:
                                    for row in table:
                                        # Join each cell in the row
                                        table_text += " ".join([str(cell) for cell in row if cell is not None]) + " "
                                        
                                # 3. Combine them to ensure NOTHING is missed
                                combined_text = page_text + " \n " + table_text
                                clean_text = re.sub(r'\s+', ' ', combined_text)
                                no_space_text = re.sub(r'\s+', '', combined_text)
                                
                                # Extract Programme
                                name_without_ext = os.path.splitext(file_name)[0]
                                if '_' in name_without_ext:
                                    programme = name_without_ext.split('_', 1)[0].strip()
                                else:
                                    programme = name_without_ext.strip()
                                    
                                # Smart Subject Extraction
                                subject_match = re.search(r'Class 6\s+(.*?)\s+Class 7', clean_text, re.IGNORECASE)
                                if subject_match:
                                    subject = subject_match.group(1).replace('|', '').strip()
                                elif '_' in name_without_ext:
                                    subject = name_without_ext.split('_', 1)[1].strip()
                                else:
                                    subject = "Unknown Subject"
                                    
                                # THE ULTIMATE DATE REGEX
                                # Finds any 1-2 digit day, 1-2 digit month, and 4 digit year (202x)
                                # Completely ignores ANY invisible garbage symbols between the numbers
                                raw_dates = re.findall(r'(\d{1,2})[^a-zA-Z0-9]*(\d{1,2})[^a-zA-Z0-9]*(202\d)', no_space_text)
                                dates = []
                                for d, m, y in raw_dates:
                                    # Format nicely with leading zeros (e.g. 05/09/2026)
                                    formatted_date = f"{int(d):02d}/{int(m):02d}/{y}"
                                    if formatted_date not in dates:
                                        dates.append(formatted_date)
                                
                                link_match = re.search(r'https?://[^\s]+', clean_text)
                                zoom_link = link_match.group(0) if link_match else ""
                                
                                day_match = re.search(r'\b(Monday|Tuesday|Wednesday|Thursday|Friday|Saturday|Sunday)\b', clean_text)
                                day = day_match.group(0) if day_match else ""
                                
                                time_match = re.search(r'\d{1,2}:\d{2}\s*[AP]M\s*-\s*\d{1,2}:\d{2}\s*[AP]M', clean_text, re.IGNORECASE)
                                if time_match:
                                    time_str = time_match.group(0)
                                else:
                                    times = re.findall(r'\d{1,2}:\d{2}', clean_text)
                                    if len(times) >= 2:
                                        t_ints = [int(t.split(':')[0])*60 + int(t.split(':')[1]) for t in times[:2]]
                                        sorted_times = [x for _, x in sorted(zip(t_ints, times[:2]))]
                                        time_str = f"{sorted_times[0]} - {sorted_times[1]} PM"
                                    else:
                                        time_str = ""
                                        
                                # Build the Rows
                                if dates:
                                    for i, date_val in enumerate(dates):
                                        all_rows.append({
                                            'Programme & Semester': programme,
                                            'Subject Name': subject,
                                            'Class': f"Class {i+1}",
                                            'Day': day,
                                            'Date': date_val,
                                            'Time (PM)': time_str,
                                            'Zoom Link': zoom_link
                                        })
                                else:
                                    st.warning(f"⚠️ Skipped Page {page_num + 1}. No dates found.")
                                    # THE DEBUG EXPANDER
                                    with st.expander(f"🔍 Click to view EXACT RAW TEXT for Page {page_num + 1}"):
                                        st.write("If the dates are missing from the box below, it means the PDF is structurally damaged or locked.")
                                        st.code(combined_text)
                                        
                            except Exception as e:
                                st.warning(f"⚠️ Error on Page {page_num + 1} of '{file_name}': {e}")
                                
                except Exception as e:
                    st.error(f"❌ Critical error opening {file_name}: {e}")
        
        # 6. Format and Output
        if all_rows:
            df = pd.DataFrame(all_rows)
            st.success(f"✅ Successfully processed {len(all_rows)} total classes!")
            
            output = BytesIO()
            df.to_excel(output, index=False, sheet_name="Schedule")
            output.seek(0)
            
            wb = openpyxl.load_workbook(output)
            ws = wb.active
            
            merge_cols = [1, 2, 4, 6, 7]
            for col_idx in merge_cols:
                start_row = 2
                for row_idx in range(3, ws.max_row + 2):
                    val_current = ws.cell(row=row_idx, column=col_idx).value if row_idx <= ws.max_row else None
                    val_start = ws.cell(row=start_row, column=col_idx).value
                    
                    if val_current != val_start:
                        if row_idx - 1 > start_row:
                            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=row_idx - 1, end_column=col_idx)
                            ws.cell(row=start_row, column=col_idx).alignment = Alignment(vertical='center', horizontal='center')
                        start_row = row_idx
            
            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)
            
            st.download_button(
                label="⬇️ Download Formatted Excel Sheet",
                data=final_output,
                file_name='Consolidated_Live_Session_Time_Table.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
        else:
            st.error("No valid data could be extracted. Please check your PDFs.")
